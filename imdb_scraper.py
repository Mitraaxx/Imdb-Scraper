import requests
from bs4 import BeautifulSoup
import argparse
from urllib.parse import quote_plus
import time
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def get_best_parser():
    try:
        import lxml
        return 'lxml'
    except ImportError:
        print("Note: lxml is not installed. Using standard HTML parser instead.")
        print("For better performance, consider installing lxml via:")
        print("pip install lxml==4.9.3")
        return 'html.parser'

class IMDBSeriesScraper:
    def __init__(self, excel_file="imdb_series_data.xlsx"):
        self.base_url = "https://www.imdb.com"
        self.search_url = "https://www.imdb.com/find?q="
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Referer': 'https://www.imdb.com/'
        }
        self.excel_file = excel_file
        self.html_parser = get_best_parser()

    def search_series(self, query):
        """Search for a series on IMDb and return potential matches"""
        try:
            print(f"\nSearching IMDb for: '{query}'...")
            search_response = requests.get(
                f"{self.search_url}{quote_plus(query)}&s=tt&ttype=tv",
                headers=self.headers,
                timeout=10
            )
            search_response.raise_for_status()
            
            soup = BeautifulSoup(search_response.text, self.html_parser)
            results = []
            
            result_items = soup.find_all('li', class_='ipc-metadata-list-summary-item')
            
            for item in result_items[:5]:  
                title_element = item.find('a', class_='ipc-metadata-list-summary-item__t')
                if not title_element:
                    continue
                    
                title = title_element.text.strip()
                link = title_element['href'].split('?')[0]
                results.append({
                    'title': title,
                    'url': f"{self.base_url}{link}"
                })
                
            return results if results else None
            
        except Exception as e:
            print(f"Search error: {str(e)}")
            return None

    def get_series_rating(self, series_url):
        """Get the rating for a specific series"""
        try:
            print(f"Fetching details from: {series_url}")
            time.sleep(1)  
            series_response = requests.get(series_url, headers=self.headers, timeout=10)
            series_response.raise_for_status()
            
            soup = BeautifulSoup(series_response.text, self.html_parser)
            
            rating_element = soup.select_one('[data-testid="hero-rating-bar__aggregate-rating__score"] span:first-child')
            rating = rating_element.text.strip() if rating_element else "N/A"
            
            title_element = soup.select_one('h1[data-testid="hero__pageTitle"] span')
            title = title_element.text if title_element else "Unknown Title"
            
            genre_elements = soup.select('div.ipc-chip-list__scroller span.ipc-chip__text')
            genres = ", ".join([genre.text for genre in genre_elements[:3]]) if genre_elements else "N/A"
            
            return {
                'title': title,
                'rating': rating,
                'url': series_url,
                'genres': genres
            }
            
        except Exception as e:
            print(f"Rating fetch error: {str(e)}")
            return None

    def get_series_rating_by_name(self, series_name):
        """Search for a series by name and return its rating"""
        search_results = self.search_series(series_name)
        
        if not search_results:
            print(f"\n❌ No results found for '{series_name}'")
            return None
            
        print(f"\n🔍 Found {len(search_results)} results:")
        for i, result in enumerate(search_results, 1):
            print(f"{i}. {result['title']}")
        
        print("\nFetching rating for the first result...")
        return self.get_series_rating(search_results[0]['url'])
    
    def save_to_excel(self, data):
        """Save data to Excel file, appending if file exists"""
        try:
            if os.path.exists(self.excel_file):
                wb = load_workbook(self.excel_file)
                ws = wb.active
                
                # Check if this series already exists in the file
                existing_urls = [row[3].value for row in ws.iter_rows(min_row=2, max_col=4)]
                if data['url'] in existing_urls:
                    print(f"ℹ️ Series '{data['title']}' already exists in the Excel file")
                    return False
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "IMDb TV Series"
                
                # Define headers
                headers = ["Title", "Rating", "Genres", "URL"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Add new data
            new_row = [
                data['title'],
                data['rating'],
                data['genres'],
                data['url']
            ]
            ws.append(new_row)
            
            # Apply borders to new row
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for col_num in range(1, 5):
                ws.cell(row=ws.max_row, column=col_num).border = border
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = (max_length + 2) if max_length < 50 else 50
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(self.excel_file)
            print(f"✅ Data successfully saved to '{self.excel_file}'")
            return True
            
        except Exception as e:
            print(f"\n❌ Error saving Excel file: {str(e)}")
            return False


def main():
    parser = argparse.ArgumentParser(description="IMDb TV Series Rating Checker")
    parser.add_argument('series', nargs='+', help="Name of the TV series to search")
    parser.add_argument('--excel', '-e', help="Specify Excel filename for output (optional)")
    parser.add_argument('--multiple', '-m', action='store_true', help="Search for multiple series separated by commas")
    args = parser.parse_args()
    
    excel_file = args.excel if args.excel else "imdb_series_data.xlsx"
    scraper = IMDBSeriesScraper(excel_file)
    
    if args.multiple:
        series_list = ' '.join(args.series).split(',')
        series_list = [s.strip() for s in series_list]
        
        print(f"Searching for {len(series_list)} series...")
        for series_name in series_list:
            result = scraper.get_series_rating_by_name(series_name)
            if result:
                print("\n" + "="*50)
                print(f"📺 Title: {result['title']}")
                print(f"🎭 Genres: {result['genres']}")
                print(f"⭐ Rating: {result['rating']}/10")
                print(f"🔗 URL: {result['url']}")
                print("="*50 + "\n")
                
                # Save each result immediately
                scraper.save_to_excel(result)
    else:
        series_name = ' '.join(args.series)
        result = scraper.get_series_rating_by_name(series_name)
        
        if result:
            print("\n" + "="*50)
            print(f"📺 Title: {result['title']}")
            print(f"🎭 Genres: {result['genres']}")
            print(f"⭐ Rating: {result['rating']}/10")
            print(f"🔗 URL: {result['url']}")
            print("="*50 + "\n")
            
            # Save the result
            scraper.save_to_excel(result)

if __name__ == "__main__":
    main()
    input("Press Enter to exit...")